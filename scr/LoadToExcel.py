import openpyxl as px
import io
import urllib3

from utils import *
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Side, Border, Alignment, Font

from GetMangaInfo import Manga

# Styles
fill = PatternFill("solid", fgColor="D9E1F2")
thin = Side(border_style="thin", color="000000")
border = Border(right=thin, left=thin, top=thin, bottom=thin)
aline = Alignment(horizontal="center", vertical="center")

name_font = Font(name="Calibri", size=14, bold=True)
name_aline = Alignment(horizontal="left", vertical="center")

genre_font = Font(name="Calibri", size=16, bold=True)

count_font = Font(name="Calibri", size=20, bold=True)
count_font_lauft = Font(name="Calibri", size=20, bold=True, color='FF0000')

finished_font = Font(name="Calibri", size=22, bold=True)

def add_to_excel_file(path, manga: Manga, manga_have_count: int):
    # Loading the excel file and catching errors
    print(f"\nLädt '{path}'...")
    try:
        wb = px.load_workbook(path)
        print("Datai 'path' wurde erfolgreich geladen!\n")
    except FileNotFoundError:
        print("Datei wurde nicht gefunden!\n")
        input("Drücke 'Enter' um zurückzukehren...")
        return False

    sheet = wb.active

    # Starting from the next empty row beginning at row 4
    for i, row in enumerate(sheet['B']):
        if i < 4:
            continue

        if row.value is None:
            cur = str(i + 1)
            break

        if i == len(sheet['B']) - 1:
            cur = str(len(sheet['B']) + 1)

    # Loading the cover into the 'A' columne
    if manga.cover:
        http = urllib3.PoolManager()
        req = http.request('GET', manga.cover)
        image_file = io.BytesIO(req.data)
        img = Image(image_file)

        img.anchor = "A" + cur
        img.width = 96
        img.height = 134
        sheet.add_image(img, "A" + cur)

    sheet.row_dimensions[int(cur)].height = 100

    # Loading the name into the 'B' columne
    name_cell = "B" + cur

    sheet[name_cell].font = name_font
    sheet[name_cell].alignment = name_aline
    sheet[name_cell].fill = fill
    sheet[name_cell].border = border
    sheet[name_cell] = manga.name
    sheet[name_cell].hyperlink = manga.link

    # Loading the author into the 'C' columne
    author_cell = "C" + cur

    sheet[author_cell].font = name_font
    sheet[author_cell].alignment = aline
    sheet[author_cell].fill = fill
    sheet[author_cell].border = border
    sheet[author_cell] = manga.author

    # Loading the genre into the 'D' columne
    genre_cell = "D" + cur

    sheet[genre_cell].font = genre_font
    sheet[genre_cell].alignment = aline
    sheet[genre_cell].fill = fill
    sheet[genre_cell].border = border
    sheet[genre_cell] = manga.genre

    # Loading the have count into the 'E' columne
    count_cell = "E" + cur

    sheet[count_cell].font = count_font
    sheet[count_cell].alignment = aline
    sheet[count_cell].fill = fill
    sheet[count_cell].border = border
    sheet[count_cell] = manga_have_count

    # Loading the german- and max count into the 'F' columne
    counts_cell = "F" + cur

    sheet[counts_cell].font = count_font_lauft if not manga.finished else count_font
    sheet[counts_cell].alignment = aline
    sheet[counts_cell].fill = fill
    sheet[counts_cell].border = border
    sheet[counts_cell].number_format = "@"
    if manga.max_count == manga.german_count:
        sheet[counts_cell] = manga.max_count
    else:
        sheet[counts_cell] = str(manga.german_count) + "/" + str(manga.max_count)

    # Loading the cost into the 'G' columne
    cost_cell = "G" + cur

    sheet[cost_cell].font = count_font
    sheet[cost_cell].alignment = aline
    sheet[cost_cell].fill = fill
    sheet[cost_cell].border = border
    sheet[cost_cell] = manga.cost
    sheet[cost_cell].number_format = "0.00€"

    wb.save(path)

    print(f"Der Manga '{manga.name}' wurde erfolgreich zur Liste hinzugefügt!\n")

    return True